import json
import os
import pickle
import shutil
import uuid
from datetime import datetime

import openpyxl
import uvicorn
from fastapi import FastAPI, WebSocket
from llama_index.core import Settings, SimpleDirectoryReader, StorageContext, KnowledgeGraphIndex
from llama_index.core.graph_stores import SimpleGraphStore
from openpyxl.utils import get_column_letter
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse

from api.article_writing import get_outline, get_summary, get_keywords, extract_content_from_json, get_body, \
    list_to_query, revise_article, get_outline_by_shanhuyun, shanhuyun_get_body
from database.graph_ngql import create_nebula_space_and_schema, drop_space
from database.sql import get_user_with_menus, check_username_exists, insert_user, insert_knowledge, \
    get_knowledge_by_user, delete_knowledge_by_name_and_user, insert_history_qa, query_history_by_user_and_type, \
    insert_knowledge_file, query_knowledge_file_by_knowledge_id_and_file_name, delete_knowledge_file_by_name_and_file, \
    update_user
from knowledge.dataset_api import matching_paragraph
from llm.embeddings import bg3_m3, rerank

from llm.glm4 import glm4_9b_chat_ws, glm4_9b_chat_long_http, glm4_9b_chat_ws_common
from llm.glm4_llamaindex import GLM4
from llm.llama_index_embeddings import InstructorEmbeddings
from llm.zhipuai_llamaindex import BianCangLLM
from milvus.milvus_tools import create_milvus, insert_milvus, get_milvus_collections_info, del_entity, \
    get_unique_field_values, delete_milvus, del_entity_by_file, query_milvus_by_file_name
from prompt import file_chat_prompt
from prompts import del_japanese_prompt, del_japanese_prompt_ws
from util.websocket_utils import ConnectionManager
from utils import md5_encrypt, download_file, put_file, has_japanese, get_red_text_from_docx, \
    replace_text_in_docx, parse_file_other, parse_file_pdf, matching_milvus_paragraph, get_graph
from models.entity import Question, UserLogin, UserRegister, Basic, Article, Edit, JachatCorrect, \
    JafileCorrect, Filechat1, Filechat2, ResponseEntity, Knowledge, GetKnowledge, DelKnowledge, KnowledgeQa, \
    KnowledgeFile, KnowledgeFileDel, KnowledgeFileUpload, HistoryList, KnowledgeFileList


def init_flask():
    Settings.embed_model = InstructorEmbeddings()
    Settings.llm = BianCangLLM()

    app = FastAPI()
    #前端跨域 添加 CORS 中间件
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],  # 允许所有来源，或者指定特定的源列表
        allow_credentials=True,
        allow_methods=["*"],  # 允许所有方法（如GET、POST等）
        allow_headers=["*"],  # 允许所有请求头
    )

    # 登录
    @app.post("/api/user/login")
    def login(user: UserLogin):
        print(user)
        # 获取用户及其菜单信息
        result = get_user_with_menus(user.userid, md5_encrypt(user.password), user.language)

        if not result:
            return JSONResponse(status_code=500, content={"message": "用户名或密码不存在"})

        # 提取用户信息和菜单信息
        user_info = result["user_info"]
        menu_list = result["menuList"]
        tokenstr = user_info[0] + user.userid + user.password
        # 生成Token
        token = md5_encrypt(tokenstr)
        # 构建返回的数据结构
        return JSONResponse(content={
            "userId": user_info[0],
            "userName": user_info[1],
            "roles": user_info[2],
            "desc": user_info[3],
            "password": user_info[4],
            "email": user_info[5],
            "iphone": user_info[6],
            "menuList": menu_list,
            "token": token
        })

    # 注册
    @app.post("/api/user/register")
    def register(user: UserRegister):
        # 获取用户及密码
        user_id = user.userid
        user_name = user.username
        user_password = md5_encrypt(user.password)
        # 判断用户名是否存在
        if check_username_exists(user_id):
            return JSONResponse(status_code=200, content={"message": "用户名存在"})
        else:
            insert_user(user_id, user_name, '用户', '1,2,3,4,5', '', user_password, user.email,
                        user.iphone)
            return JSONResponse(status_code=200, content={"message": "success"})

    # 修改个人信息
    @app.post("/api/user/update_userinfo")
    def update_userinfo(user: UserRegister):
        try:
            user_id = user.userid
            user_name = user.username
            if user.password == '':
                user_password = ''
            else:
                user_password = md5_encrypt(user.password)
            update_user(user_id, user_name, user_password, user.email, user.iphone)
            return JSONResponse(status_code=200, content={"message": "success"})
        except Exception as e:
            return JSONResponse(status_code=500, content={"message": e})

    @app.post("/api/history/get_list")
    def get_history_list(history: HistoryList):
        try:
            res = query_history_by_user_and_type(history.user_id, history.type)
            history = []
            for item in res:
                data = [
                    {"role": "user", "content": item[0]},
                    {"role": "assistant", "content": item[1]},
                    {"time": item[2]}
                ]
                history.append(data)
            return ResponseEntity(
                message=history,
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=str(e),
                status_code=500
            )

    # 普通对话
    @app.websocket("/api/common_chat/{v1}")
    async def common_chat(websocket: WebSocket, v1: str):

        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = Question.parse_raw(data)
            print(params)
            prompt = params.prompt
            history = params.history
            history.append({"content": prompt, "role": "user"})
            temperature = params.temperature
            aaa = ""
            answer = glm4_9b_chat_ws_common(history, temperature)
            for chunk in answer:
                aaa += str(chunk.choices[0].delta.content)
                print(chunk.choices[0].delta.content, end="")
                await manager.send_personal_message(json.dumps({
                    "answer": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
            await manager.send_personal_message(json.dumps({
                "type": "stop"
            }, ensure_ascii=False), websocket)
            insert_history_qa(v1, prompt, aaa, "dialog")
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 论文---获取关键词
    @app.post("/api/write/get_keywords")
    def get_write_keywords():
        try:
            res = get_unique_field_values("damage_explosion_v2", "type")
            keywords_list = [{"name": item, "checked": False} for item in res]
            return ResponseEntity(
                message=keywords_list,
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=str(e),
                status_code=500
            )

    # 论文---获取关键词
    @app.post("/api/write/get_keywords_shunhuyun")
    def get_write_keywords_shanhuyun():
        try:
            res = get_unique_field_values("jianting", "type")
            keywords_list = [{"name": item, "checked": False} for item in res]
            return ResponseEntity(
                message=keywords_list,
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=str(e),
                status_code=500
            )

    # 论文---生成大纲
    @app.post("/api/write/get_basic")
    def get_basic(basic: Basic):
        print(basic)
        #     return JSONResponse(status_code=200, content={
        #   "摘要": "我是摘要",
        #   "标题": "第五章 结论与展望",
        #   "内容": "总结研究成果，并对未来的研究方向进行展望。",
        #   "小节": [
        #     {
        #       "小节标题": "5.1 研究结论",
        #       "内容": "概括研究的主要发现和结论"
        #     },
        #     {
        #       "小节标题": "5.2 研究局限",
        #       "内容": "讨论本研究的局限性和未来改进的空间"
        #     },
        #     {
        #       "小节标题": "5.3 研究展望",
        #       "内容": "展望未来游戏测试研究的可能方向和领域"
        #     }
        #   ],
        #   "yinyong": "我是引用",
        # })
        return get_outline(basic.article_title, 0.7, list_to_query(basic.article_choices))

    # 论文---生成论文
    @app.websocket("/api/write/get_article/{v1}")
    async def get_article(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = Article.parse_raw(data)
            article_base = params.article_base
            article_choices = params.article_choices
            ref_file = []
            #摘要
            await manager.send_personal_message(json.dumps({
                "title": article_base['标题']
            }, ensure_ascii=False), websocket)
            summary = get_summary(article_base, list_to_query(article_choices))
            for chunk in summary["ai_say"]:
                await manager.send_personal_message(json.dumps({
                    "summary": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
            #关键字
            keywords = get_keywords(article_base)
            await manager.send_personal_message(json.dumps({
                "keywords": keywords
            }, ensure_ascii=False), websocket)
            # 正文
            json_list = extract_content_from_json(article_base)
            for index, item in enumerate(json_list):
                print(index, item)
                if '标题' in item and index != len(json_list) - 1:
                    await manager.send_personal_message(json.dumps({
                        "biaoti": item['标题']
                    }, ensure_ascii=False), websocket)
                elif '小节标题' in item:
                    await manager.send_personal_message(json.dumps({
                        "xiaojiebiaoti": item['小节标题']
                    }, ensure_ascii=False), websocket)
                    body = get_body(article_base, str(item), list_to_query(article_choices))
                    if body['ref_file']:
                        for x in body['ref_file']:
                            ref_file.append(x)
                    for chunk in body['ai_say']:
                        await manager.send_personal_message(json.dumps({
                            "xiaojieneirong": chunk.choices[0].delta.content
                        }, ensure_ascii=False), websocket)
                elif '标题' in item and index == len(json_list) - 1:
                    await manager.send_personal_message(json.dumps({
                        "xiaojiebiaoti": item['标题']
                    }, ensure_ascii=False), websocket)
                    body = get_body(article_base, str(item), list_to_query(article_choices))
                    for chunk in body['ai_say']:
                        await manager.send_personal_message(json.dumps({
                            "xiaojieneirong": chunk.choices[0].delta.content
                        }, ensure_ascii=False), websocket)
                    if body['ref_file']:
                        for x in body['ref_file']:
                            ref_file.append(x)
            # 引用
            await manager.send_personal_message(json.dumps({
                "yinyong": ref_file
            }, ensure_ascii=False), websocket)
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 论文---修改章节
    @app.websocket("/api/write/edit_article/{v1}")
    async def edit_article(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = Edit.parse_raw(data)
            oldpart = params.oldpart
            prompt = params.prompt
            revise_text = revise_article(oldpart, prompt)
            for chunk in revise_text:
                await manager.send_personal_message(json.dumps({
                    "answer": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
                print(chunk.choices[0].delta.content, end="")
            await manager.send_personal_message(json.dumps({
                "answer": 'stop_finished'
            }, ensure_ascii=False), websocket)
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 日语修正-对话
    @app.websocket("/api/correctJa/chat/{v1}")
    async def correct_ja_chat(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = JachatCorrect.parse_raw(data)
            prompt = params.prompt
            answer = del_japanese_prompt_ws(prompt)
            ai_say = ""
            for chunk in answer:
                ai_say += chunk.choices[0].delta.content
                await manager.send_personal_message(json.dumps({
                    "answer": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
            await manager.send_personal_message(json.dumps({
                "type": "stop"
            }, ensure_ascii=False), websocket)
            insert_history_qa(v1, prompt, ai_say, "revise")
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 日语修正-文件
    @app.websocket("/api/correctJa/file/{v1}")
    async def correct_ja_file(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = JafileCorrect.parse_raw(data)
            bucket_name = params.bucket_name
            object_name = params.object_name
            download_file_res = download_file(bucket_name, object_name)
            file_path = download_file_res['file_path']
            file_dir = download_file_res['file_dir']
            file_type = object_name.split('.')[1]
            new_file_path = file_path.replace(object_name, 'new_' + object_name)
            if file_type == 'xlsx' or file_type == 'xls':
                # 加载工作簿和工作表
                workbook = openpyxl.load_workbook(file_path)
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    processed = False
                    # 遍历每个单元格
                    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
                        for cell in row:
                            if isinstance(cell.value, str) and has_japanese(cell.value):
                                # 处理日文内容
                                new_value = del_japanese_prompt(0.1, cell.value, [])
                                processed = True
                                cell_value = str(cell.value).replace("\n", "")
                                new_cell_value = str(new_value).replace("\n", "")
                                await manager.send_personal_message(json.dumps({
                                    "content": f'''修正：
                                    - **シート**: {sheet} ，**行**: {cell.row} ，**列**: {get_column_letter(cell.column)}
                                    - **修正前**: {cell_value}
                                    - **修正後**: {new_cell_value}
                                '''
                                }, ensure_ascii=False), websocket)
                                insert_history_qa(v1, "日语修正-excel", f'''修正：
                                    - **シート**: {sheet} ，**行**: {cell.row} ，**列**: {get_column_letter(cell.column)}
                                    - **修正前**: {cell_value}
                                    - **修正後**: {new_cell_value}
                                ''', "revise")
                                cell.value = new_value
                    # 如果工作表被处理，保存修改
                    if processed:
                        workbook.save(new_file_path)
            if file_type == 'txt':
                # 读取文件内容
                with open(file_path, 'r', encoding='utf-8') as file:
                    text = file.read()
                text_array = text.split("\n\n")
                text_array = [item for item in text_array if item.strip()]
                for index, item in enumerate(text_array):
                    if has_japanese(item):
                        ai_value = del_japanese_prompt(0.1, item, [])
                        aisay_item = item.replace("\n", "")
                        aisay_value = ai_value.replace("\n", "")
                        text_array[index] = ai_value  # 替换原始文本数组中的值
                        await manager.send_personal_message(json.dumps({
                            "content": f'''修正：
                            - **修正前**: {aisay_item}
                            - **修正後**: {aisay_value}
                           '''
                        }, ensure_ascii=False), websocket)
                        insert_history_qa(v1, "日语修正-txt", f'''修正：
                            - **修正前**: {aisay_item}
                            - **修正後**: {aisay_value}''', "revise")
                # 将修改后的文本数组重新组合成字符串
                new_text = '\n\n'.join(text_array)
                # 指定一个绝对路径
                output_path = new_file_path
                # 创建一个txt文件，写入new_text
                with open(output_path, 'w', encoding='utf-8') as output_file:
                    output_file.write(new_text)
            if file_type == 'doc' or file_type == 'docx':
                replacements = {}
                red_list = get_red_text_from_docx(file_path)
                if len(red_list) > 0:
                    for index, item in enumerate(red_list):
                        ai_value = del_japanese_prompt(0.1, item, [])
                        aisay_item = item.replace("\n", "")
                        aisay_value = ai_value.replace("\n", "")
                        replacements[item] = ai_value
                        if len(aisay_item) > 5:
                            await manager.send_personal_message(json.dumps({
                                "content": f'''修正：
                                - **修正前**: {aisay_item}
                                - **修正後**: {aisay_value}
                            '''
                            }, ensure_ascii=False), websocket)
                            insert_history_qa(v1, "日语修正-word", f'''修正：
                            - **修正前**: {aisay_item}
                            - **修正後**: {aisay_value}
                            ''', "revise")
                        for key in list(
                                replacements.keys()):  # Using list to avoid runtime error due to dictionary size change
                            if len(key) < 5:
                                replacements[key] = ""
                        replace_text_in_docx(file_path, replacements, new_file_path)
                else:
                    await manager.send_personal_message(json.dumps({
                        "content": f'''翻訳が必要なテキストは検出されませんでした'''
                    }, ensure_ascii=False), websocket)
            put_file("modify-ja-file", f"{file_dir}.{file_type}", new_file_path)
            shutil.rmtree(f"./data/{file_dir}")
            await manager.send_personal_message(json.dumps({
                "old_file_name": object_name,
                "new_file_name": f"{file_dir}.{file_type}",
                "bucket_name": "modify-ja-file"
            }, ensure_ascii=False), websocket)
            insert_history_qa(v1, "修改后-下载文件", f"192.168.2.8:19000/modify-ja-file/{file_dir}.{file_type}",
                              "revise")
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 文件对话-上传文件
    @app.post("/api/file_chat/upload")
    def file_chat_upload(file: Filechat1):
        try:
            result = download_file(file.bucket_name, file.object_name)
            file_dir = result['file_dir']
            documents = SimpleDirectoryReader(f'./data/{file_dir}/').load_data()

            graph_store = SimpleGraphStore()

            storage_context = StorageContext.from_defaults(graph_store=graph_store)

            filechat_index = KnowledgeGraphIndex.from_documents(documents=documents,
                                                                max_triplets_per_chunk=3,
                                                                storage_context=storage_context,
                                                                include_embeddings=True)
            # 将 filechat_index 序列化并保存到文件
            with open("filechat_index.pkl", "wb") as f:
                pickle.dump(filechat_index, f)
            return ResponseEntity(
                message="success",
                status_code=200
            )
        except Exception as e:
            print(e)
            return ResponseEntity(
                message="error",
                status_code=500
            )

    # 文件对话-用户对话
    @app.websocket("/api/file_chat/qa/{v1}")
    async def file_chat_qa(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)

        try:
            data = await websocket.receive_text()
            params = Filechat2.parse_raw(data)
            question = params.question
            language = params.language
            question = f'''{question}。请默认用{language}回答问题。'''
            with open("filechat_index.pkl", "rb") as f:
                filechat_index = pickle.load(f)

            query_engine = filechat_index.as_query_engine(include_text=True,
                                                          response_mode="tree_summarize",
                                                          embedding_mode="hybrid",
                                                          similarity_top_k=4,
                                                          streaming=True)
            generator = query_engine.query(question).response_gen
            ai_say = ""
            for chunk in generator:
                # 在这里处理每个块
                ai_say += chunk
                await manager.send_personal_message(json.dumps({
                    "answer": chunk
                }, ensure_ascii=False), websocket)
            # history = params.history
            # user_id = "_" + params.userid
            # language = params.language
            # res = matching_paragraph(question, user_id, 1000)
            # filtered_result = []
            # for result in res:
            #     filtered_result = [res.entity.text for res in result if res.score > 0]
            # a = rerank(question, filtered_result, 3)
            # rerank_results = [y['index'] for y in a if y['relevance_score'] > 0.7]
            # rerank_filtered_result = []
            # for index in rerank_results:
            #     rerank_filtered_result.append(filtered_result[index])
            # message = {"content": file_chat_prompt.format(question=question, content=str(rerank_filtered_result),
            #                                               language=language), "role": "user"}
            # history.append(message)
            # answer = glm4_9b_chat_ws(history, 0.7)
            # ai_say = ""
            # for chunk in answer:
            #     ai_say += chunk.choices[0].delta.content
            #     await manager.send_personal_message(json.dumps({
            #         "answer": chunk.choices[0].delta.content
            #     }, ensure_ascii=False), websocket)
            # await manager.send_personal_message(json.dumps({
            #     "type": "stop"
            # }, ensure_ascii=False), websocket)
            insert_history_qa(v1, question, ai_say, "fileDialog")
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 知识库-获取知识库列表
    @app.post("/api/knowledge/get_list")
    def get_knowledge_list(knowledge: GetKnowledge):
        keys = ['id', 'name', 'description', 'milvus_name', 'graph_name', 'user_id', 'create_time']
        knowledge_list = get_knowledge_by_user(knowledge.userid)
        result = []
        for item in knowledge_list:
            result.append(dict(zip(keys, item)))
        return ResponseEntity(
            message=result,
            status_code=200
        )

    # 知识库-新建知识库
    @app.post("/api/knowledge/create_database")
    def create_knowledge(knowledge: Knowledge):
        try:
            res = get_knowledge_by_user(knowledge.userid)
            second_elements = [item[1] for item in res]
            if knowledge.name in second_elements:
                return ResponseEntity(
                    message="知识库名称已存在！",
                    status_code=501
                )
            knowledge_name = knowledge.name + "_" + knowledge.userid
            # 创建milvus
            create_milvus(knowledge_name, knowledge.description)
            # 创建graph
            create_nebula_space_and_schema(knowledge_name)
            # 插入mysql
            insert_knowledge(str(uuid.uuid4()), knowledge.name, knowledge.description, knowledge_name, knowledge_name,
                             knowledge.userid, datetime.now())
            return ResponseEntity(
                message="success",
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message="error",
                status_code=500
            )

    # 知识库-删除知识库
    @app.post("/api/knowledge/drop_database")
    def drop_knowledge(knowledge: DelKnowledge):
        try:
            mg_name = knowledge.name + "_" + knowledge.userid
            drop_space(mg_name)
            delete_milvus(mg_name)
            delete_knowledge_by_name_and_user(knowledge.name, knowledge.userid)
            return ResponseEntity(
                message="success",
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message="error",
                status_code=500
            )

    # 对话-知识库
    @app.websocket("/api/knowledge/qa/{v1}")
    async def knowledge_qa(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = KnowledgeQa.parse_raw(data)
            history = params.history
            question = params.question
            knowledge_name = params.knowledge_name
            language = params.language
            user_id = params.userid
            ai_say = ""
            # if knowledge_name == "xiyouji_1":
            #     if '孙悟空是一个什么样的人' in question:
            #         question = '孙悟空是一个什么样的人，请有调理的梳理'
            #     query_engine = index.as_query_engine(include_text=True,
            #                                          response_mode="tree_summarize",
            #                                          embedding_mode="hybrid",
            #                                          similarity_top_k=4,
            #                                          streaming=True)
            #     generator = query_engine.query(question).response_gen
            #     for chunk in generator:
            #         # 在这里处理每个块
            #         ai_say += chunk
            #         await manager.send_personal_message(json.dumps({
            #             "answer": chunk
            #         }, ensure_ascii=False), websocket)
            # else:
            res = matching_milvus_paragraph(question, knowledge_name, 3)
            messages = {"content": file_chat_prompt.format(question=question, content=str(res), language=language),
                        "role": "user"}
            history.append(messages)
            answer = glm4_9b_chat_ws(history, 0.1)
            for chunk in answer:
                ai_say += chunk.choices[0].delta.content
                await manager.send_personal_message(json.dumps({
                    "answer": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
            print(ai_say)
            await manager.send_personal_message(json.dumps({
                "type": "stop"
            }, ensure_ascii=False), websocket)
            insert_history_qa(v1, question, ai_say, "questions")
        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    # 知识库-获取知识库内文件列表
    @app.post("/api/knowledge/get_files")
    def get_files(knowledge: KnowledgeFile):
        try:
            res = get_unique_field_values(knowledge.knowledge_name + "_" + knowledge.user_id, "file_name")
            return ResponseEntity(
                message=res,
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=e,
                status_code=500
            )

    # 知识库-知识库内文件列表删除
    @app.post("/api/knowledge/del_file")
    def del_file(knowledge: KnowledgeFileDel):
        try:
            knowledge_name = knowledge.knowledge_name + "_" + knowledge.user_id
            del_entity_by_file(knowledge_name, knowledge.file_name)
            delete_knowledge_file_by_name_and_file(knowledge_name, knowledge.file_name)
            return ResponseEntity(
                message="success",
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=e,
                status_code=500
            )

    # 知识库-知识库内文件上传
    @app.post("/api/knowledge/upload_file")
    def upload_file(knowledge: KnowledgeFileUpload):
        try:
            knowledge_name = knowledge.knowledge_name + "_" + knowledge.user_id
            file_name = knowledge.minio_file_name
            bucket_name = knowledge.minio_bucket_name
            file_type = os.path.splitext(file_name)[1]
            if file_type == '.pdf':
                text_splitter = parse_file_pdf(bucket_name, file_name)
            else:
                text_splitter = parse_file_other(bucket_name, file_name)
            for text in text_splitter:
                data = [{
                    'text': text,
                    'file_name': knowledge.file_name,
                    'embeddings': bg3_m3(text)
                }]
                insert_milvus(data, knowledge_name)
            prompt = f"""
            {text_splitter}\n\t
            请你帮我总结上述文本中的大纲，结果不要超过100字，只返回大纲，不需要多余的解释。
            """
            message = [{"content": prompt, "role": "user"}]
            graph_html = get_graph(bucket_name, file_name)
            outline = glm4_9b_chat_long_http(message, 0.2)
            insert_knowledge_file(knowledge_name, knowledge.file_name, graph_html, outline)
            return ResponseEntity(
                message="success",
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=e,
                status_code=500
            )

    # 知识库-文件内列表
    @app.post("/api/knowledge/file_list")
    def get_file_list(knowledge: KnowledgeFileList):
        try:
            knowledge_name = knowledge.knowledge_name + "_" + knowledge.user_id
            file_name = knowledge.file_name
            text_splitter = query_milvus_by_file_name(knowledge_name, file_name)
            graph = query_knowledge_file_by_knowledge_id_and_file_name(knowledge_name, file_name)
            res = {
                "text_splitter": text_splitter,
                "graph": graph
            }
            return ResponseEntity(
                message=res,
                status_code=200
            )
        except Exception as e:
            return ResponseEntity(
                message=e,
                status_code=500
            )

    # 珊瑚云_大纲接口
    @app.post("/api/write/get_basic_shanhuyun")
    def get_basic(basic: Basic):
        print(basic)
        return get_outline_by_shanhuyun(basic.article_title, 0.7, list_to_query(basic.article_choices))

    # 珊瑚云_生成科技研究报告
    @app.websocket("/api/write/get_article_shanhuyun/{v1}")
    async def get_article(websocket: WebSocket, v1: str):
        manager = ConnectionManager()
        await manager.connect(websocket)
        try:
            data = await websocket.receive_text()
            params = Article.parse_raw(data)
            article_base = params.article_base
            article_choices = params.article_choices
            ref_file = []
            await manager.send_personal_message(json.dumps({
                "title": article_base['标题']
            }, ensure_ascii=False), websocket)
            #摘要
            summary = get_summary(article_base, list_to_query(article_choices))
            #关键字
            keywords = get_keywords(article_base)
            for chunk in summary["ai_say"]:
                print(chunk.choices[0].delta.content, end='')
                await manager.send_personal_message(json.dumps({
                    "summary": chunk.choices[0].delta.content
                }, ensure_ascii=False), websocket)
            await manager.send_personal_message(json.dumps({
                "keywords": keywords
            }, ensure_ascii=False), websocket)
            # 正文
            json_list = extract_content_from_json(article_base)
            for index, item in enumerate(json_list):
                print(index, item)
                if '标题' in item and index != len(json_list) - 1:
                    await manager.send_personal_message(json.dumps({
                        "biaoti": item['标题']
                    }, ensure_ascii=False), websocket)
                elif '小节标题' in item:
                    await manager.send_personal_message(json.dumps({
                        "xiaojiebiaoti": item['小节标题']
                    }, ensure_ascii=False), websocket)
                    body = shanhuyun_get_body(article_base, str(item), list_to_query(article_choices))
                    if body['ref_file']:
                        for x in body['ref_file']:
                            ref_file.append(x)
                    for chunk in body['ai_say']:
                        print(chunk.choices[0].delta.content, end='')
                        await manager.send_personal_message(json.dumps({
                            "xiaojieneirong": chunk.choices[0].delta.content
                        }, ensure_ascii=False), websocket)
                elif '标题' in item and index == len(json_list) - 1:
                    await manager.send_personal_message(json.dumps({
                        "xiaojiebiaoti": item['标题']
                    }, ensure_ascii=False), websocket)
                    body = shanhuyun_get_body(article_base, str(item), list_to_query(article_choices))
                    for chunk in body['ai_say']:
                        print(chunk.choices[0].delta.content, end='')
                        await manager.send_personal_message(json.dumps({
                            "xiaojieneirong": chunk.choices[0].delta.content
                        }, ensure_ascii=False), websocket)
                    if body['ref_file']:
                        for x in body['ref_file']:
                            ref_file.append(x)
            # 引用
            await manager.send_personal_message(json.dumps({
                "yinyong": ref_file
            }, ensure_ascii=False), websocket)

        except Exception as e:
            print(e)
        finally:
            manager.disconnect(websocket)

    return app


app = init_flask()
uvicorn.run(app, host='0.0.0.0', port=8009, workers=1)
